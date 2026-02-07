import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict
import os
import logging

from sport import config

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def fetch_data(page: int = 0) -> Dict:
    """
    Fetch data from the API for the given page number.
    """
    url = f'https://api.worldarchery.sport/?EventTypeId=1&WorldRecordStatus=0&WorldRankingEvent=0&Cancelled=-1&Detailed=1&StartDate=2026-01-01&EndDate=2026-12-30&SortBy=DATE&v=3&content=COMPETITIONS&RBP=100000'
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7,la;q=0.6,und;q=0.5',
        'Client-Security-Token': '8a9aeaee-69b6-4d7b-81fa-8ca3f93adb6c',
        'Connection': 'keep-alive',
        'DNT': '1',
        'Origin': 'https://www.worldarchery.sport',
        'Referer': 'https://www.worldarchery.sport/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"'
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # Raise an error for bad responses
        logger.info(f"Successfully fetched data")
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching data")
        raise


def save_to_excel(competitions: List[Dict], filename):
    """
    Save competition data to Excel file with auto-adjusted column widths.
    """
    logger.info(f"Saving data to Excel file: {filename}")

    # Flatten the data for better Excel representation
    flattened_data = []
    for comp in competitions:
        flat_comp = {
            'ID': comp['ID'],
            'Name': comp['Name'],
            'Venue': comp['Venue'],
            'Place': comp['Place'],
            'Country': comp['Country'],
            'CountryName': comp['CountryName'],
            'DFrom': comp['DFrom'],
            'DTo': comp['DTo'],
            'Level': comp['Level'],
            'SubLevel': comp['SubLevel'],
            'IsFeatured': comp['IsFeatured'],
            'WorldRecordStatus': comp['WorldRecordStatus'],
            'WorldRankingEvent': comp['WorldRankingEvent'],
            'IsLive': comp['IsLive'],
            'IsCancelled': comp['IsCancelled'],
            'StatusDescription': comp['StatusDescription'],
            'Address': comp.get('Address', {}).get('Address', ''),
            'Latitude': comp.get('Address', {}).get('Latitude', ''),
            'Longitude': comp.get('Address', {}).get('Longitude', '')
        }
        # 只保留开始时间、结束时间都在2026年的比赛
        if flat_comp['DFrom'].startswith('2026') and flat_comp['DTo'].startswith('2026'):
            flattened_data.append(flat_comp)

    # Convert to DataFrame
    df = pd.DataFrame(flattened_data)

    # Save to Excel
    try:
        df.to_excel(filename, index=False, sheet_name='Competitions')
        # Load workbook for formatting
        wb = load_workbook(filename)
        ws = wb.active

        # Auto-adjust column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Find the maximum length of content in each column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Add some padding to the maximum length
            adjusted_width = (max_length + 2)

            # Set column width
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the formatted workbook
        wb.save(filename)
        logger.info(f"Data saved to {filename} with auto-adjusted column widths")
    except Exception as e:
        logger.error(f"Error saving data to Excel: {e}")
        raise


def main():
    # Initialize variables
    all_competitions = []
    page = 0

    logger.info("Starting data fetch and save process.")

    while True:
        # Fetch data for the current page
        data = fetch_data(page=page)

        # Extract competition items
        competitions = data.get('items', [])
        all_competitions.extend(competitions)

        # Check if we've reached the last page
        total_results = data.get('pageInfo', {}).get('totalResults', 0)
        results_per_page = data.get('pageInfo', {}).get('resultsPerPage', 20)
        logger.info(f"Fetched {len(competitions)} competitions, total results: {total_results}")
        if len(competitions) < results_per_page or len(all_competitions) >= total_results:
            break

        # Increment the page number for the next request
        page += 1

    # Save all competitions to Excel
    filename = os.path.basename(__file__)
    basename = os.path.splitext(filename)[0]
    new_filename = basename + ".xlsx"
    excel_file_path = os.path.join(config.output_path, new_filename)

    save_to_excel(all_competitions, excel_file_path)


if __name__ == "__main__":
    main()