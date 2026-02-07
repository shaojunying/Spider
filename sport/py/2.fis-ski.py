import logging
import os
from typing import List, Dict

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from sport import config


def fetch_fis_calendar(season_code: str, season_month: str) -> str:
    """
    Fetch FIS calendar data from the website.

    Args:
        season_code: Season code (e.g., '2024')
        season_month: Month to fetch data for (format: 'MM-YYYY')

    Returns:
        HTML content of the calendar
    """
    print("参数：", season_code, season_month)
    # url = 'https://data.fis-ski.com/fis_events/ajax/calendarfunctions/load_calendar.html?sectorcode=&seasoncode=&categorycode=&disciplinecode=&gendercode=&place=&eventselection=&racedate=&racecodex=&nationcode=&seasonmonth=02-2026&loadmonth=1&saveselection=-1&pageid=725&seasonselection='
    url = (f'https://data.fis-ski.com/fis_events/ajax/calendarfunctions/load_calendar.html?'
           f'seasoncode={season_code}&seasonmonth={season_month}&loadmonth=0'
           f'&saveselection=-1&pageid=725')

    print(f'Fetching FIS calendar data from {url}')

    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return None


def parse_fis_competition_data(html_content: str) -> List[Dict]:
    """
    Parse FIS competition data from HTML content and return structured data.

    Args:
        html_content: HTML string containing FIS competition data

    Returns:
        List of dictionaries containing parsed competition information
    """
    if not html_content:
        return []

    soup = BeautifulSoup(html_content, 'html.parser')
    competitions = []

    rows = soup.find_all('div', class_='table-row')

    for row in rows:
        if not row.get('id'):
            continue

        competition = {
            'event_id': row.get('id'),
            'dates': '',
            'location': '',
            'country': '',
            'discipline': '',
            'event_type': '',
            'races': '',
            'gender': set(),
            'status': {
                'data_available': False,
                'pdf_available': False,
                'changes': False,
                'cancelled': False
            }
        }

        status_items = row.find_all('span', class_='status__item')
        for item in status_items:
            if 'status__item_selected' in item.get('class', []):
                title = item.get('title', '').lower()
                if 'data available' in title:
                    competition['status']['data_available'] = True
                elif 'pdf available' in title:
                    competition['status']['pdf_available'] = True

        date_element = row.find('a', class_='g-lg-4')
        if date_element:
            competition['dates'] = date_element.text.strip()

        location_element = row.find('div', class_='clip bold font_md_large font_lg_large')
        if location_element:
            competition['location'] = location_element.text.strip()

        country_element = row.find('span', class_='country__name-short')
        if country_element:
            competition['country'] = country_element.text.strip()

        discipline_element = row.find('a', class_='g-lg-1 g-md-1 hidden-sm-down')
        if discipline_element and discipline_element.text.strip():
            competition['discipline'] = discipline_element.text.strip()

        event_info = row.find('div', class_='split-row split-row_bordered')
        if event_info:
            info_items = event_info.find_all('div', class_='clip')
            if len(info_items) >= 1:
                competition['event_type'] = info_items[0].text.strip()
            if len(info_items) >= 2:
                competition['races'] = info_items[1].text.strip()

        gender_container = row.find('a', class_='g-lg-1 g-md-1 g-sm-2 hidden-sm-down bold')
        if gender_container:
            gender_items = gender_container.find_all('div', class_='gender__item')
            for item in gender_items:
                if 'gender__item_l' in item.get('class', []) and item.text.strip() == 'W':
                    competition['gender'].add('Women')
                elif 'gender__item_m' in item.get('class', []) and item.text.strip() == 'M':
                    competition['gender'].add('Men')

        competition['gender'] = ' & '.join(sorted(list(competition['gender'])))
        competitions.append(competition)

    return competitions


def save_to_excel(competitions: List[Dict], filename):
    """
    Save competition data to Excel file with auto-adjusted column widths.

    Args:
        competitions: List of competition dictionaries
        filename: Name of the Excel file to save
    """
    # Flatten the status dictionary for better Excel representation
    flattened_data = []
    for comp in competitions:
        flat_comp = comp.copy()
        for status_key, status_value in comp['status'].items():
            flat_comp[f'status_{status_key}'] = status_value
        del flat_comp['status']
        flattened_data.append(flat_comp)

    # Convert to DataFrame
    df = pd.DataFrame(flattened_data)

    # Reorder columns for better readability
    column_order = [
        'event_id', 'dates', 'location', 'country', 'discipline',
        'event_type', 'races', 'gender', 'status_data_available',
        'status_pdf_available', 'status_changes', 'status_cancelled'
    ]
    df = df[column_order]

    # Rename columns for better readability
    column_names = {
        'status_data_available': 'Data Available',
        'status_pdf_available': 'PDF Available',
        'status_changes': 'Changes',
        'status_cancelled': 'Cancelled'
    }
    df = df.rename(columns=column_names)

    # Save to Excel
    df.to_excel(filename, index=False, sheet_name='FIS Competitions')

    # Load workbook for formatting
    wb = load_workbook(filename)
    ws = wb.active

    # Auto-adjust column widths based on content
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # Find the maximum length of content in each column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Add some padding to the maximum length
        adjusted_width = (max_length + 2)

        # Set column width
        ws.column_dimensions[column_letter].width = adjusted_width
    print(filename)
    # Save the formatted workbook
    wb.save(filename)
    print(f"Data saved to {filename} with auto-adjusted column widths")


def main():
    competitions = []
    for month in range(1, 13):
        print(f"Fetching data for month {month}")
        html_content = fetch_fis_calendar(season_code='2026', season_month=f'{month:02d}-2026')

        if html_content:
            data = parse_fis_competition_data(html_content)
            # 判断两个列表中是否存在完全一样的元素
            for item in data:
                if item not in competitions:
                    competitions.append(item)
            print("len(competitions):", len(competitions))
        else:
            print("Failed to fetch data")

    filename = os.path.basename(__file__)
    basename = os.path.splitext(filename)[0]
    new_filename = basename + ".xlsx"
    excel_file_path = os.path.join(config.output_path, new_filename)
    # Save to Excel file
    save_to_excel(competitions, excel_file_path)


if __name__ == "__main__":
    main()